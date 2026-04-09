import hashlib
import logging
import json
import os
from datetime import datetime, timezone

import openpyxl
from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)

OUTPUT_DIR = os.getenv("OUTPUT_DIR", "./output")


def _generate_chunk_id(
    file_path: str,
    sheet_name: str,
    row_num: int
) -> str:
    """Generate a unique chunk ID using SHA256 hash."""
    return hashlib.sha256(
        (file_path + sheet_name + str(row_num)).encode("utf-8")
    ).hexdigest()[:16]


def _save_blocks(blocks: list[dict]) -> None:
    """Append blocks to scraped.jsonl output file."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, "scraped.jsonl")
    try:
        with open(output_path, "a", encoding="utf-8") as f:
            for block in blocks:
                f.write(
                    json.dumps(block, ensure_ascii=False) + "\n"
                )
        logger.info(f"Saved {len(blocks)} blocks to {output_path}")
    except Exception as e:
        logger.error(f"Failed to save blocks: {e}")


def load_excel(
    file_path    : str,
    document_type: str,
    org_name     : str
) -> list[dict]:
    """
    Load and extract text from an Excel file.
    Each row becomes one block.
    Returns list of block dicts.
    """
    logger.info(f"Loading Excel: {file_path}")

    if not os.path.exists(file_path):
        logger.error(f"Excel file not found: {file_path}")
        return []

    document_title = os.path.basename(file_path)
    extracted_at   = datetime.now(timezone.utc).isoformat()
    blocks         = []

    try:
        workbook = openpyxl.load_workbook(
            file_path, read_only=True, data_only=True
        )
    except Exception as e:
        logger.error(f"Failed to open Excel file {file_path}: {e}")
        return []

    for sheet_name in workbook.sheetnames:
        logger.info(f"Processing sheet: {sheet_name}")

        try:
            sheet = workbook[sheet_name]
            rows  = list(sheet.iter_rows(values_only=True))
        except Exception as e:
            logger.warning(
                f"Failed to read sheet {sheet_name}: {e}"
            )
            continue

        if not rows:
            logger.debug(f"Sheet {sheet_name} is empty — skipping")
            continue

        # First row = headers
        headers = [
            str(cell).strip() if cell is not None else f"Column{i}"
            for i, cell in enumerate(rows[0], start=1)
        ]

        # Each subsequent row = one block
        for row_num, row in enumerate(rows[1:], start=2):
            # Build text as "Header: value, Header: value"
            row_parts = []
            for header, cell in zip(headers, row):
                value = str(cell).strip() if cell is not None else ""
                if value and value.lower() != "none":
                    row_parts.append(f"{header}: {value}")

            if not row_parts:
                logger.debug(
                    f"Skipping empty row {row_num} "
                    f"in sheet {sheet_name}"
                )
                continue

            text     = ", ".join(row_parts)
            chunk_id = _generate_chunk_id(
                file_path, sheet_name, row_num
            )

            block = {
                "chunk_id"      : chunk_id,
                "source_url"    : file_path,
                "document_title": document_title,
                "organization"  : org_name,
                "section_title" : sheet_name,
                "section_level" : "",
                "chapter"       : "",
                "article"       : "",
                "content_type"  : "paragraph",
                "text"          : text,
                "char_count"    : len(text),
                "extracted_at"  : extracted_at,
            }
            blocks.append(block)

    workbook.close()
    _save_blocks(blocks)

    logger.info(
        f"Excel loading complete. "
        f"Rows extracted: {len(blocks)} from {file_path}"
    )
    return blocks